"""
vibecoding.ipynb Step 1(취합)·Step 2(검증) 코드 셀을 그대로 옮긴 모듈.

목적은 이미 검증된 노트북 코드를 재사용하는 것이므로 로직을 바꾸지 않았다.
노트북과 다른 점은 딱 두 가지뿐이다:
  1. sample_root/master_path를 함수 인자로 받는다(노트북은 셀 상수를 그대로 참조했음).
  2. 셀 맨 아래에서 즉시 실행하던 `run_merge_pipeline()` / `run_validation_pipeline()`
     호출을 지웠다 — 언제 실행할지는 웹앱(app.py)이 결정한다.
print() 로그는 그대로 두었다. app.py는 이 print 출력을 stdout 캡처로 그대로 가져와
화면에 보여준다(노트북 셀 출력을 보는 것과 동일한 경험).
"""

import csv as csv_module
import glob
import os
import re
from io import BytesIO

import matplotlib
import openpyxl
from dotenv import load_dotenv
from openai import OpenAI

matplotlib.use("Agg")
import matplotlib.pyplot as plt

# 노트북 셀은 load_dotenv() 호출 없이 os.environ.get(...)만 썼다 — Jupyter를 이미
# .env가 로드된 셸에서 띄웠을 때만 우연히 동작했을 가능성이 높다. 독립 실행되는
# 웹앱에서도 확실히 동작하도록 여기서 명시적으로 .env를 읽어들인다.
load_dotenv()

SAMPLE_ROOT = "example/input_data"
MASTER_FILE = "example/result/2026년도_예산집행취합.xlsx"
EMAIL_DIR = "example/result/이메일"
HEADER = ["사업장", "부서명", "예산과목", "배정액", "집행액", "작성자", "비고"]

FOLDER_PATTERN = re.compile(r"^(\d{4})(\d{2})_")
DATA_SHEET_PATTERN = re.compile(r"^\d{4}-\d{2}$")

EXPECTED_DEPTS = [
    ("본원", "재무회계부"),
    ("본원", "디지털정보부"),
    ("본원", "안전경영부"),
    ("우주", "활동협력부"),
    ("평창", "활동협력부"),
]


# ===========================================================================
# Step 1 — 취합 (vibecoding.ipynb 코드 셀 그대로)
# ===========================================================================

def ensure_master_file(master_path):
    """마스터 파일이 없으면 새로 만든다."""
    if os.path.exists(master_path):
        print(f"마스터 파일 확인: {master_path}")
    else:
        os.makedirs(os.path.dirname(master_path), exist_ok=True)
        wb = openpyxl.Workbook()
        wb.save(master_path)
        print(f"마스터 파일이 없어 새로 생성했습니다: {master_path}")


def parse_sheet_name(folder_name):
    match = FOLDER_PATTERN.match(folder_name)
    if not match:
        return None
    year, month = match.groups()
    return f"{year}-{month}"


def find_unmerged_folders(sample_root, master_path):
    wb = openpyxl.load_workbook(master_path)
    existing_sheets = set(wb.sheetnames)

    targets = []
    for name in sorted(os.listdir(sample_root)):
        folder_path = os.path.join(sample_root, name)
        if not os.path.isdir(folder_path):
            continue

        sheet_name = parse_sheet_name(name)
        if sheet_name is None:
            print(f"건너뜀 (이름 규칙 불일치): {name}")
            continue

        if sheet_name in existing_sheets:
            print(f"건너뜀 (이미 취합됨): {name} → {sheet_name} 시트")
        else:
            print(f"신규 발견: {name} → {sheet_name} 시트로 취합 예정")
            targets.append((folder_path, sheet_name))

    return targets


def load_department_files(input_dir):
    """폴더 안의 부서 제출 엑셀을 읽어 합친다. 'original'로 시작하는 보관용 원본 파일은 제외한다."""
    merged_rows = []
    all_paths = sorted(glob.glob(os.path.join(input_dir, "*.xlsx")))

    file_paths = [p for p in all_paths if not os.path.basename(p).lower().startswith("original")]
    skipped = [p for p in all_paths if p not in file_paths]
    for p in skipped:
        print(f"  제외 (보관용 원본): {os.path.basename(p)}")

    for path in file_paths:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        file_rows = list(ws.iter_rows(min_row=2, values_only=True))
        merged_rows.extend(file_rows)
        print(f"  읽음: {os.path.basename(path)} ({len(file_rows)}행)")

    if skipped:
        print(f"  → original 파일 {len(skipped)}개 제외됨")

    return merged_rows


def _is_blank_default_sheet(ws):
    return (
        ws.title == "Sheet"
        and ws.max_row == 1
        and ws.max_column == 1
        and ws["A1"].value is None
    )


def save_to_master(rows, master_path, sheet_name):
    wb = openpyxl.load_workbook(master_path)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.append(HEADER)
    for row in rows:
        ws.append(row)

    if "Sheet" in wb.sheetnames and wb["Sheet"] is not ws and _is_blank_default_sheet(wb["Sheet"]):
        del wb["Sheet"]

    wb.save(master_path)
    print(f"  저장 완료: {master_path} ({sheet_name} 시트, {len(rows)}행)")


def run_merge_pipeline(sample_root=SAMPLE_ROOT, master_path=MASTER_FILE):
    ensure_master_file(master_path)
    targets = find_unmerged_folders(sample_root, master_path)

    for folder_path, sheet_name in targets:
        print(f"\n[{sheet_name}] 취합 시작: {folder_path}")
        rows = load_department_files(folder_path)
        save_to_master(rows, master_path, sheet_name)


# ===========================================================================
# Step 2 — 검증 (vibecoding.ipynb 코드 셀 그대로)
# ===========================================================================

def find_unvalidated_sheets(master_path):
    wb = openpyxl.load_workbook(master_path)
    existing_sheets = set(wb.sheetnames)

    targets = []
    for name in wb.sheetnames:
        if not DATA_SHEET_PATTERN.match(name):
            continue

        log_sheet = f"검증로그_{name}"
        if log_sheet in existing_sheets:
            print(f"건너뜀 (이미 검증됨): {name} → {log_sheet} 존재")
        else:
            print(f"신규 발견: {name} → 검증 필요")
            targets.append(name)

    return targets


def load_merged_rows(master_path, sheet_name):
    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb[sheet_name]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return [dict(zip(header, values)) for values in ws.iter_rows(min_row=2, values_only=True)]


def validate(rows, expected_depts):
    submitted_depts = {(r["사업장"], r["부서명"]) for r in rows}
    missing_depts = [d for d in expected_depts if d not in submitted_depts]

    missing_rows = [r for r in rows if r["집행액"] is None]

    outlier_rows = [
        r for r in rows
        if r["집행액"] is not None and r["집행액"] > r["배정액"]
    ]

    return missing_depts, missing_rows, outlier_rows


def print_report(sheet_name, missing_depts, missing_rows, outlier_rows):
    print(f"\n=== [{sheet_name}] 검증 결과 ===")

    print(f"[미제출 부서] {len(missing_depts)}건")
    for 사업장, 부서명 in missing_depts:
        print(f"  - {사업장} {부서명}")

    print(f"[집행액 누락] {len(missing_rows)}건")
    for r in missing_rows:
        print(f"  - {r['사업장']} {r['부서명']} / {r['예산과목']} (작성자: {r['작성자']})")

    print(f"[이상치(집행액 > 배정액)] {len(outlier_rows)}건")
    for r in outlier_rows:
        print(
            f"  - {r['사업장']} {r['부서명']} / {r['예산과목']} : "
            f"배정액 {r['배정액']:,}원 < 집행액 {r['집행액']:,}원 (작성자: {r['작성자']})"
        )


def save_validation_log(master_path, sheet_name, missing_depts, missing_rows, outlier_rows):
    log_sheet = f"검증로그_{sheet_name}"
    wb = openpyxl.load_workbook(master_path)

    if log_sheet in wb.sheetnames:
        del wb[log_sheet]
    ws = wb.create_sheet(log_sheet)
    ws.append(["구분", "내용"])

    for 사업장, 부서명 in missing_depts:
        ws.append(["미제출부서", f"{사업장} {부서명}"])

    for r in missing_rows:
        ws.append(["누락값", f"{r['사업장']} {r['부서명']} / {r['예산과목']} 집행액 미기재"])

    for r in outlier_rows:
        ws.append([
            "이상치",
            f"{r['사업장']} {r['부서명']} / {r['예산과목']} "
            f"배정액 {r['배정액']:,}원 < 집행액 {r['집행액']:,}원",
        ])

    wb.save(master_path)
    print(f"저장 완료: {master_path} ({log_sheet} 시트)")


def run_validation_pipeline(master_path=MASTER_FILE):
    targets = find_unvalidated_sheets(master_path)

    for sheet_name in targets:
        rows = load_merged_rows(master_path, sheet_name)
        missing_depts, missing_rows, outlier_rows = validate(rows, EXPECTED_DEPTS)
        print_report(sheet_name, missing_depts, missing_rows, outlier_rows)
        save_validation_log(master_path, sheet_name, missing_depts, missing_rows, outlier_rows)


# ===========================================================================
# Step 3-1 — 확인 요청 이메일 초안 (vibecoding.ipynb 코드 셀 그대로, LLM · Upstage API)
# ===========================================================================

_upstage_client = OpenAI(
    api_key=os.environ.get("UPSTAGE_API_KEY"),
    base_url="https://api.upstage.ai/v1",
)


def find_months_to_process(master_path, email_dir):
    """검증로그_연도-월 시트가 있는 모든 달을 찾는다. 재실행 시 기존 이메일도 재생성한다."""
    wb = openpyxl.load_workbook(master_path)

    targets = []
    for name in wb.sheetnames:
        if not name.startswith("검증로그_"):
            continue
        sheet_name = name[len("검증로그_"):]

        email_path = os.path.join(email_dir, f"이메일_{sheet_name}.txt")
        if os.path.exists(email_path):
            print(f"재생성: {sheet_name} → {email_path} 덮어쓰기")
        else:
            print(f"신규 발견: {sheet_name} → 이메일 생성 필요")
        targets.append(sheet_name)

    return targets


def load_log_rows(master_path, sheet_name):
    """검증로그_연도-월 시트를 읽어 딕셔너리 리스트로 반환한다."""
    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb[f"검증로그_{sheet_name}"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return [dict(zip(header, values)) for values in ws.iter_rows(min_row=2, values_only=True)]


def load_담당자_map(master_path, sheet_name):
    """(사업장, 부서명) → 작성자 매핑을 원본 취합 시트에서 가져온다."""
    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb[sheet_name]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    담당자맵 = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(header, values))
        담당자맵[(row["사업장"], row["부서명"])] = row["작성자"]
    return 담당자맵


def group_issues_by_dept(log_rows):
    """검증로그의 누락값/이상치 행을 부서별로 묶는다."""
    grouped = {}
    for r in log_rows:
        if r["구분"] not in ("누락값", "이상치"):
            continue
        match = re.match(r"^(\S+)\s+(\S+)\s*/\s*(.+)$", r["내용"])
        if not match:
            continue
        사업장, 부서명, 내용 = match.groups()
        grouped.setdefault((사업장, 부서명), []).append(f"[{r['구분']}] {내용}")
    return grouped


def draft_email(사업장, 부서명, 담당자, 연도월, issues):
    """Upstage Solar Pro에게 확인 요청 이메일 초안을 요청한다."""
    이슈목록 = "\n".join(f"- {i}" for i in issues)
    prompt = f"""당신은 본원 재무회계부 홍길동입니다. 아래 부서 담당자에게 예산 집행 데이터 확인을 요청하는 이메일을 정중한 업무용 한국어로 작성해주세요.

- 받는 사람: {담당자} ({사업장} {부서명})
- 대상 기간: {연도월}
- 확인이 필요한 사항:
{이슈목록}

요청 사항:
- 제목과 본문을 구분해서 작성해주세요.
- 본문에는 각 확인 사항을 항목별로 정리하고, 사유와 함께 회신해달라고 요청해주세요.
- 정중하지만 간결하게 작성해주세요.
- "제목: ...", 그 다음 줄에 "본문:"과 내용 순서로 출력해주세요.
"""
    response = _upstage_client.chat.completions.create(
        model="solar-pro",
        messages=[{"role": "user", "content": prompt}],
    )
    return response.choices[0].message.content


def build_email_texts(master_path, sheet_name):
    log_rows = load_log_rows(master_path, sheet_name)
    grouped = group_issues_by_dept(log_rows)

    if not grouped:
        print(f"[{sheet_name}] 확인 필요 없음 (이상치/누락값 없음)")
        return []

    담당자맵 = load_담당자_map(master_path, sheet_name)

    email_blocks = []
    for (사업장, 부서명), issues in grouped.items():
        담당자 = 담당자맵.get((사업장, 부서명), "담당자")
        print(f"  이메일 작성 중: {담당자} ({사업장} {부서명}) - {len(issues)}건")
        content = draft_email(사업장, 부서명, 담당자, sheet_name, issues)
        header = f"받는사람: {담당자} ({사업장} {부서명})\n이메일: [담당자 이메일 주소를 입력하세요]\n\n"
        email_blocks.append(header + content)

    return email_blocks


def save_emails(email_dir, sheet_name, email_blocks):
    os.makedirs(email_dir, exist_ok=True)
    path = os.path.join(email_dir, f"이메일_{sheet_name}.txt")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n\n---\n\n".join(email_blocks))
    print(f"저장 완료: {path} ({len(email_blocks)}명)")


def run_email_pipeline(master_path=MASTER_FILE, email_dir=EMAIL_DIR):
    targets = find_months_to_process(master_path, email_dir)

    for sheet_name in targets:
        email_blocks = build_email_texts(master_path, sheet_name)
        if email_blocks:
            save_emails(email_dir, sheet_name, email_blocks)


# ===========================================================================
# 웹앱 전용 어댑터 — Step 1을 브라우저 디렉토리 업로드(st.file_uploader의
# accept_multiple_files="directory")로 실행하기 위한 최소한의 추가 코드.
#
# 업로드된 파일은 디스크 경로가 아니라 메모리상의 파일 객체라서 glob.glob으로
# 스캔할 수 없다 — 그래서 이 부분만 새로 작성했다. 그 외 판정 로직(폴더명 →
# 연도-월 파싱은 parse_sheet_name, 'original' 제외는 동일 규칙, 저장은
# save_to_master)은 전부 위의 노트북 코드를 그대로 재사용한다.
#
# 브라우저가 디렉토리 업로드 시 각 파일의 name에 상대경로(예:
# "202607_부서제출파일/예산집행_..._202607.xlsx")를 실어 보내주는 것을 이용해서
# 파일이 어느 하위 폴더 소속인지 판단한다.
# ===========================================================================

SUPPORTED_UPLOAD_EXTENSIONS = (".xlsx", ".csv")


def filter_supported_files(uploaded_files):
    """xlsx/csv 확장자만 남기고 나머지(.DS_Store 등)는 제외한다.

    브라우저가 디렉토리 전체를 선택하는 경우(webkitdirectory) accept/type 필터가
    적용되지 않는다 — 폴더 선택 창에는 파일 단위로 고를 UI 자체가 없기 때문에,
    브라우저가 이 제약을 그냥 무시하고 폴더 안의 모든 파일을 돌려준다. 그래서
    st.file_uploader의 type=["xlsx","csv"]만으로는 걸러지지 않고, 업로드된 목록을
    받은 뒤 여기서 다시 한번 확장자로 걸러야 한다.
    """
    kept, excluded = [], []
    for f in uploaded_files:
        if f.name.lower().endswith(SUPPORTED_UPLOAD_EXTENSIONS):
            kept.append(f)
        else:
            excluded.append(f.name)
    return kept, excluded


def group_uploads_by_folder(uploaded_files):
    """디렉토리 업로드로 받은 파일들을 상대경로의 상위 폴더 이름 기준으로 묶는다.

    반환값: (groups: {폴더이름: [업로드파일, ...]}, ungrouped: [파일이름, ...])
    """
    groups = {}
    ungrouped = []
    for f in uploaded_files:
        parts = f.name.replace("\\", "/").split("/")
        if len(parts) >= 2:
            folder_name = parts[-2]
            groups.setdefault(folder_name, []).append(f)
        else:
            ungrouped.append(f.name)
    return groups, ungrouped


def _coerce_csv_value(value):
    """CSV는 전부 문자열로 들어오므로, 엑셀에서 읽을 때와 비슷하게 숫자/빈값을 맞춰준다."""
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        pass
    try:
        return float(value)
    except (TypeError, ValueError):
        pass
    return value


def _read_rows_from_upload(f):
    """xlsx/csv 확장자에 따라 알맞게 읽어 동일한 형식(값 튜플 리스트, 헤더 제외)으로 반환한다."""
    f.seek(0)
    if f.name.lower().endswith(".csv"):
        text = f.read().decode("utf-8-sig")
        rows = list(csv_module.reader(text.splitlines()))
        return [tuple(_coerce_csv_value(v) for v in row) for row in rows[1:]]

    wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
    ws = wb.active
    return list(ws.iter_rows(min_row=2, values_only=True))


def load_department_files_from_uploads(files):
    """load_department_files()와 동일한 로직 — 소스만 디스크 폴더 대신 업로드 파일 목록.
    xlsx·csv 둘 다 지원한다(그 외 형식은 업로더 자체에서 이미 걸러짐)."""
    file_objs = [f for f in files if not os.path.basename(f.name).lower().startswith("original")]
    skipped = [f for f in files if f not in file_objs]
    for f in skipped:
        print(f"  제외 (보관용 원본): {os.path.basename(f.name)}")

    merged_rows = []
    for f in file_objs:
        file_rows = _read_rows_from_upload(f)
        merged_rows.extend(file_rows)
        print(f"  읽음: {os.path.basename(f.name)} ({len(file_rows)}행)")

    if skipped:
        print(f"  → original 파일 {len(skipped)}개 제외됨")

    return merged_rows


def run_merge_pipeline_from_uploads(uploaded_files, master_path):
    """run_merge_pipeline()과 동일한 흐름 — 소스만 OS 폴더 스캔 대신 브라우저 디렉토리 업로드.
    이미 취합된 달은 그대로 건너뛴다(정정 반영은 이 함수가 아니라 아래 Step 5 쪽 담당)."""
    uploaded_files, unsupported = filter_supported_files(uploaded_files)
    if unsupported:
        print(f"건너뜀 (xlsx/csv가 아닌 파일 {len(unsupported)}개 제외): {', '.join(unsupported)}")

    ensure_master_file(master_path)

    groups, ungrouped = group_uploads_by_folder(uploaded_files)
    if ungrouped:
        print(f"건너뜀 (하위 폴더 없이 업로드된 파일 {len(ungrouped)}개): {', '.join(ungrouped)}")

    wb = openpyxl.load_workbook(master_path)
    existing_sheets = set(wb.sheetnames)

    for folder_name, files in sorted(groups.items()):
        sheet_name = parse_sheet_name(folder_name)
        if sheet_name is None:
            print(f"건너뜀 (이름 규칙 불일치): {folder_name}")
            continue
        if sheet_name in existing_sheets:
            print(f"건너뜀 (이미 취합됨): {folder_name} → {sheet_name} 시트")
            continue

        print(f"\n[{sheet_name}] 취합 시작: {folder_name}")
        rows = load_department_files_from_uploads(files)
        save_to_master(rows, master_path, sheet_name)


# ===========================================================================
# Step 5 — 정정 반영 재취합·재검증 (vibecoding.ipynb 코드 셀과 동일한 판단 로직:
# 'original' 접두어 백업 파일의 존재로 정정된 달을 자동 감지) — 업로드 어댑터.
#
# 원본 노트북은 shutil.move()로 실제 디스크의 부서 폴더에서 original_ 파일을
# example/input_data/original_data로 옮긴다. 브라우저 업로드는 사용자의 실제 절대
# 경로를 알 수 없어(브라우저가 보안상 노출하지 않음) 같은 방식으로 파일을 옮길 수
# 없다 — 대신 업로드된 바이트를 웹앱이 관리하는 보관 폴더에 그대로 저장해서 같은
# '원본 백업을 한곳에 모아 보관' 역할을 하게 했다. 그 외 감지·재취합·재검증 로직은
# find_folders_needing_reprocess/force_save_to_master/force_revalidate와 동일하다.
# ===========================================================================

def find_reprocess_groups_from_uploads(uploaded_files):
    """find_folders_needing_reprocess()와 동일한 판단 — 업로드된 파일 중 'original'로
    시작하는 xlsx가 포함된 월별 폴더(=사람이 정정한 달)를 찾는다."""
    uploaded_files, _ = filter_supported_files(uploaded_files)
    groups, _ = group_uploads_by_folder(uploaded_files)

    targets = []
    for folder_name, files in sorted(groups.items()):
        sheet_name = parse_sheet_name(folder_name)
        if sheet_name is None:
            continue
        original_files = [
            f for f in files
            if os.path.basename(f.name).lower().startswith("original")
            and f.name.lower().endswith(".xlsx")
        ]
        if original_files:
            print(f"수정 감지: {folder_name} → {sheet_name} (original 파일 {len(original_files)}개)")
            targets.append((folder_name, sheet_name, files, original_files))

    return targets


def archive_original_uploads(original_files, archive_dir):
    """archive_original_files()와 같은 역할 — original 파일을 보관 폴더로 옮긴다.
    (업로드 바이트는 실제 디스크 경로가 없어 shutil.move 대신 그대로 저장하는 방식으로 구현)"""
    os.makedirs(archive_dir, exist_ok=True)
    for f in original_files:
        f.seek(0)
        dst = os.path.join(archive_dir, os.path.basename(f.name))
        with open(dst, "wb") as out:
            out.write(f.read())
        print(f"  [1/3 보관 이동] {os.path.basename(f.name)} → {archive_dir}/")


def run_reprocess_pipeline_from_uploads(uploaded_files, master_path):
    """run_reprocess_pipeline()과 동일한 흐름 — 소스만 OS 폴더 스캔 대신 브라우저 디렉토리
    업로드. 'original' 백업 파일이 있는 달을 사람이 지정하지 않아도 자동으로 찾아
    재취합·재검증한다."""
    archive_dir = os.path.join(os.path.dirname(master_path) or ".", "original_data")
    targets = find_reprocess_groups_from_uploads(uploaded_files)

    if not targets:
        print("재처리할 항목이 없습니다.")
        return []

    reprocessed = []
    for folder_name, sheet_name, files, original_files in targets:
        print(f"\n=== [{sheet_name}] 재처리 시작 ===")
        archive_original_uploads(original_files, archive_dir)
        # load_department_files_from_uploads()가 'original' 파일은 이미 제외하므로
        # 남은(=정정된) 파일들만 자연히 다시 읽힌다.
        rows = load_department_files_from_uploads(files)
        save_to_master(rows, master_path, sheet_name)
        revalidate_month(master_path, sheet_name)
        reprocessed.append(sheet_name)

    return reprocessed


# ===========================================================================
# 웹앱 전용 헬퍼 — 노트북에는 없던, 화면 표시를 위한 최소한의 추가 함수.
# 취합/검증 판정 로직 자체는 건드리지 않고, 이미 위 함수들이 계산한 것을 다시 읽어
# 화면에 보여주기 좋은 형태로만 가공한다.
# ===========================================================================


def list_data_months(master_path):
    """마스터 파일에 이미 취합된 연도-월 목록(정렬됨). 사이드바 표시용."""
    if not os.path.exists(master_path):
        return []
    wb = openpyxl.load_workbook(master_path)
    return sorted(n for n in wb.sheetnames if DATA_SHEET_PATTERN.match(n))


def list_metric_months(master_path):
    """마스터 파일에 이미 지표_연도-월 시트가 있는 연도-월 목록(정렬됨). 화면 표시용."""
    if not os.path.exists(master_path):
        return []
    wb = openpyxl.load_workbook(master_path)
    return sorted(m.group(1) for name in wb.sheetnames if (m := METRIC_SHEET_PATTERN.match(name)))


def revalidate_month(master_path, sheet_name):
    """force_revalidate()와 동일 — 이미 검증된 달이라도 강제로 다시 검증하고 검증로그를
    덮어쓴 뒤, 재검증 결과 이상 유무를 요약해서 출력한다. validate()/save_validation_log()
    등은 전부 위의 노트북 코드를 그대로 재사용하고, '이미 검증됨' 게이트만 건너뛴다
    (Step 5 재처리 파이프라인에서 방금 덮어쓴 달을 즉시 다시 검증하기 위한 용도)."""
    rows = load_merged_rows(master_path, sheet_name)
    missing_depts, missing_rows, outlier_rows = validate(rows, EXPECTED_DEPTS)
    print_report(sheet_name, missing_depts, missing_rows, outlier_rows)
    save_validation_log(master_path, sheet_name, missing_depts, missing_rows, outlier_rows)

    if not missing_depts and not missing_rows and not outlier_rows:
        print(f"  ✅ [{sheet_name}] 재검증 결과 이상 없음 — 정정이 정상 반영되었습니다.")
    else:
        print(f"  ⚠️ [{sheet_name}] 재검증 결과 여전히 확인이 필요한 항목이 있습니다.")


def get_all_validation_results(master_path):
    """이미 취합된 모든 달에 대해 validate()를 다시 돌려 구조화된 결과를 얻는다.
    (파일 저장 없음, 순수 조회 — 이슈 화면에 표시할 부서/예산과목/담당자 등 구조화된
    필드가 필요한데, 검증로그 시트에는 사람이 읽기 좋은 문자열로만 저장되어 있어서
    같은 validate() 함수를 다시 호출해 원본 필드를 그대로 얻는 방식을 택했다.)
    """
    results = []
    for month in list_data_months(master_path):
        rows = load_merged_rows(master_path, month)
        missing_depts, missing_rows, outlier_rows = validate(rows, EXPECTED_DEPTS)
        results.append({
            "month": month,
            "missing_depts": missing_depts,
            "missing_rows": missing_rows,
            "outlier_rows": outlier_rows,
        })
    return results


# ===========================================================================
# Step 6 — 부서별 집행 지표 계산 (vibecoding.ipynb 코드 셀 그대로, Rule 기반)
#
# 노트북 셀은 load_merged_rows()를 Step 2와 동일한 몸체로 다시 정의하는데, 한
# 모듈에 합치면서 중복 정의 없이 위의 load_merged_rows()를 그대로 재사용했다.
# print_report()는 이름이 Step 2의 print_report()와 겹쳐서(용도가 다름)
# print_metrics_report()로만 이름을 바꿨다 — 내용은 노트북 그대로다.
# ===========================================================================

METRIC_HEADER = ["사업장", "부서명", "예산과목", "배정액", "집행액", "집행률(%)", "전월집행액", "전월대비증감액"]


def find_uncalculated_sheets(master_path):
    wb = openpyxl.load_workbook(master_path)
    existing_sheets = set(wb.sheetnames)

    targets = []
    for name in wb.sheetnames:
        if not DATA_SHEET_PATTERN.match(name):
            continue
        metric_sheet = f"지표_{name}"
        if metric_sheet in existing_sheets:
            print(f"건너뜀 (이미 계산됨): {name} → {metric_sheet} 존재")
        else:
            print(f"신규 발견: {name} → 지표 계산 필요")
            targets.append(name)

    return targets


def previous_month(sheet_name):
    year, month = map(int, sheet_name.split("-"))
    if month == 1:
        return f"{year - 1}-12"
    return f"{year}-{month - 1:02d}"


def calc_metrics(master_path, sheet_name):
    rows = load_merged_rows(master_path, sheet_name)

    wb = openpyxl.load_workbook(master_path)
    prev_sheet = previous_month(sheet_name)
    prev_map = {}
    if prev_sheet in wb.sheetnames:
        for r in load_merged_rows(master_path, prev_sheet):
            key = (r["사업장"], r["부서명"], r["예산과목"])
            prev_map[key] = r["집행액"]

    results = []
    for r in rows:
        배정액 = r["배정액"]
        집행액 = r["집행액"]
        집행률 = round(집행액 / 배정액 * 100, 1) if (집행액 is not None and 배정액) else None

        key = (r["사업장"], r["부서명"], r["예산과목"])
        전월집행액 = prev_map.get(key)

        if 집행액 is None:
            증감액 = "이번달 데이터 없음"
        elif 전월집행액 is None:
            증감액 = "전월 데이터 없음"
        else:
            증감액 = 집행액 - 전월집행액

        전월집행액_표시 = 전월집행액 if 전월집행액 is not None else "전월 데이터 없음"

        results.append({
            "사업장": r["사업장"],
            "부서명": r["부서명"],
            "예산과목": r["예산과목"],
            "배정액": 배정액,
            "집행액": 집행액,
            "집행률(%)": 집행률,
            "전월집행액": 전월집행액_표시,
            "전월대비증감액": 증감액,
        })

    return results


def print_metrics_report(sheet_name, results):
    print(f"\n=== [{sheet_name}] 지표 계산 결과 ===")
    for r in results:
        집행률_str = f"{r['집행률(%)']}%" if r["집행률(%)"] is not None else "-"
        증감 = r["전월대비증감액"]
        증감_str = f"{증감:+,}원" if isinstance(증감, (int, float)) else 증감
        print(
            f"  {r['사업장']} {r['부서명']} / {r['예산과목']}: "
            f"집행률 {집행률_str}, 전월대비 {증감_str}"
        )


def save_metrics(master_path, sheet_name, results):
    metric_sheet = f"지표_{sheet_name}"
    wb = openpyxl.load_workbook(master_path)
    if metric_sheet in wb.sheetnames:
        del wb[metric_sheet]
    ws = wb.create_sheet(metric_sheet)
    ws.append(METRIC_HEADER)
    for r in results:
        ws.append([r[col] for col in METRIC_HEADER])
    wb.save(master_path)
    print(f"저장 완료: {master_path} ({metric_sheet} 시트, {len(results)}행)")


def run_metrics_pipeline(master_path=MASTER_FILE):
    targets = find_uncalculated_sheets(master_path)
    for sheet_name in targets:
        results = calc_metrics(master_path, sheet_name)
        print_metrics_report(sheet_name, results)
        save_metrics(master_path, sheet_name, results)


# ===========================================================================
# Step 7 — 부서별 비교 그래프 생성 (vibecoding.ipynb 코드 셀 그대로, 공문 첨부용)
# ===========================================================================

GRAPH_DIR = "example/result/그래프"
METRIC_SHEET_PATTERN = re.compile(r"^지표_(\d{4}-\d{2})$")

COLOR_SURFACE = "#fcfcfb"
COLOR_PRIMARY = "#0b0b0b"
COLOR_SECONDARY = "#52514e"
COLOR_MUTED = "#898781"
COLOR_GRID = "#e1e0d9"
COLOR_BASELINE = "#c3c2b7"
COLOR_BLUE = "#2a78d6"
COLOR_RED = "#e34948"

matplotlib.rcParams["font.family"] = "AppleGothic"
matplotlib.rcParams["axes.unicode_minus"] = False


def find_months_needing_graph(master_path, graph_dir):
    """지표 시트는 있지만 그래프 파일이 아직 없는 달을 찾는다."""
    wb = openpyxl.load_workbook(master_path)
    targets = []
    for name in wb.sheetnames:
        match = METRIC_SHEET_PATTERN.match(name)
        if not match:
            continue
        sheet_name = match.group(1)
        rate_path = os.path.join(graph_dir, f"집행률_비교_{sheet_name}.png")
        if os.path.exists(rate_path):
            print(f"건너뜀 (이미 생성됨): {sheet_name} → {rate_path} 존재")
        else:
            print(f"신규 발견: {sheet_name} → 그래프 생성 필요")
            targets.append(sheet_name)
    return targets


def load_metric_rows(master_path, sheet_name):
    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb[f"지표_{sheet_name}"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return [dict(zip(header, values)) for values in ws.iter_rows(min_row=2, values_only=True)]


def aggregate_by_dept(rows):
    """예산과목 4개를 부서(사업장+부서명) 단위로 합산한다."""
    agg = {}
    for r in rows:
        key = (r["사업장"], r["부서명"])
        a = agg.setdefault(key, {"배정액": 0, "집행액": 0, "증감합": 0, "증감건수": 0, "미확정": False})
        a["배정액"] += r["배정액"] or 0
        if r["집행액"] is None:
            a["미확정"] = True
        else:
            a["집행액"] += r["집행액"]

        증감 = r["전월대비증감액"]
        if isinstance(증감, (int, float)):
            a["증감합"] += 증감
            a["증감건수"] += 1
    return agg


def _style_axes(ax):
    ax.set_facecolor(COLOR_SURFACE)
    ax.yaxis.grid(True, color=COLOR_GRID, linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    for spine in ["top", "right", "left"]:
        ax.spines[spine].set_visible(False)
    ax.spines["bottom"].set_color(COLOR_BASELINE)
    ax.tick_params(colors=COLOR_SECONDARY, labelsize=9)


def draw_rate_chart(sheet_name, agg, out_path):
    """부서별 집행률 비교 막대그래프."""
    labels = [f"{s}\n{d}" for (s, d) in agg.keys()]
    rates = [v["집행액"] / v["배정액"] * 100 if v["배정액"] else 0 for v in agg.values()]
    partial = [v["미확정"] for v in agg.values()]
    avg_rate = sum(rates) / len(rates)

    fig, ax = plt.subplots(figsize=(8, 5), facecolor=COLOR_SURFACE)
    bars = ax.bar(labels, rates, color=COLOR_BLUE, width=0.55, zorder=3)
    for b, is_partial in zip(bars, partial):
        if is_partial:
            b.set_hatch("///")
            b.set_edgecolor(COLOR_SURFACE)

    ax.axhline(avg_rate, color=COLOR_MUTED, linestyle="--", linewidth=1, zorder=2)
    ax.annotate(f"평균 {avg_rate:.1f}%", xy=(0.99, 0.95), xycoords="axes fraction",
                color=COLOR_SECONDARY, fontsize=9, ha="right", va="top")

    for b, r, is_partial in zip(bars, rates, partial):
        label = f"{r:.1f}%" + ("*" if is_partial else "")
        ax.text(b.get_x() + b.get_width() / 2, b.get_height() + max(rates) * 0.02,
                label, ha="center", fontsize=9, color=COLOR_PRIMARY)

    ax.set_title(f"{sheet_name} 부서별 예산 집행률 비교", color=COLOR_PRIMARY, fontsize=13, pad=14, loc="left")
    ax.set_ylabel("집행률 (%)", color=COLOR_SECONDARY, fontsize=10, labelpad=12)
    ax.set_ylim(0, max(rates) * 1.3)
    _style_axes(ax)

    if any(partial):
        fig.text(0.02, 0.01, "* 일부 항목 집행액 미확정 (합계에서 제외)", fontsize=8, color=COLOR_MUTED)

    fig.subplots_adjust(left=0.13, right=0.97, top=0.87, bottom=0.16)
    plt.savefig(out_path, dpi=150)
    plt.close(fig)
    print(f"  저장 완료: {out_path}")


def draw_delta_chart(sheet_name, agg, out_path):
    """부서별 전월 대비 증감 비교 막대그래프."""
    labels, deltas, partial = [], [], []
    for (s, d), v in agg.items():
        if v["증감건수"] == 0:
            continue
        labels.append(f"{s}\n{d}")
        deltas.append(v["증감합"])
        partial.append(v["미확정"])

    if not labels:
        print(f"  [{sheet_name}] 전월 비교 가능한 부서가 없어 증감 그래프를 생략합니다 (전월 데이터 없음).")
        return

    colors = [COLOR_BLUE if x >= 0 else COLOR_RED for x in deltas]

    fig, ax = plt.subplots(figsize=(8, 5), facecolor=COLOR_SURFACE)
    bars = ax.bar(labels, deltas, color=colors, width=0.55, zorder=3)
    for b, is_partial in zip(bars, partial):
        if is_partial:
            b.set_hatch("///")
            b.set_edgecolor(COLOR_SURFACE)

    span = max(deltas) - min(min(deltas), 0)
    for b, v in zip(bars, deltas):
        va = "bottom" if v >= 0 else "top"
        offset = span * 0.02 if v >= 0 else -span * 0.02
        ax.text(b.get_x() + b.get_width() / 2, v + offset, f"{v:+,.0f}",
                ha="center", fontsize=9, color=COLOR_PRIMARY, va=va)

    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    ax.axhline(0, color=COLOR_BASELINE, linewidth=1, zorder=2)
    ax.set_title(f"{sheet_name} 부서별 전월 대비 집행액 증감", color=COLOR_PRIMARY, fontsize=13, pad=14, loc="left")
    ax.set_ylabel("증감액 (원)", color=COLOR_SECONDARY, fontsize=10, labelpad=12)
    _style_axes(ax)

    if any(partial):
        fig.text(0.02, 0.01, "* 일부 항목 집행액 미확정 (합계에서 제외)", fontsize=8, color=COLOR_MUTED)

    fig.subplots_adjust(left=0.15, right=0.97, top=0.87, bottom=0.16)
    plt.savefig(out_path, dpi=150)
    plt.close(fig)
    print(f"  저장 완료: {out_path}")


def run_graph_pipeline(master_path=MASTER_FILE, graph_dir=GRAPH_DIR):
    os.makedirs(graph_dir, exist_ok=True)
    targets = find_months_needing_graph(master_path, graph_dir)

    for sheet_name in targets:
        print(f"\n[{sheet_name}] 그래프 생성 시작")
        rows = load_metric_rows(master_path, sheet_name)
        agg = aggregate_by_dept(rows)

        rate_path = os.path.join(graph_dir, f"집행률_비교_{sheet_name}.png")
        draw_rate_chart(sheet_name, agg, rate_path)

        delta_path = os.path.join(graph_dir, f"증감_비교_{sheet_name}.png")
        draw_delta_chart(sheet_name, agg, delta_path)


# ===========================================================================
# Step 8-1 — 공문 초안용 최종 자료 통합 (vibecoding.ipynb 코드 셀 그대로)
#
# 노트북 셀의 read_sheet_rows()는 load_merged_rows()와 몸체가 완전히 같아서
# 중복 정의하지 않고 load_merged_rows()를 그대로 썼다. aggregate_by_dept()도
# Step 7과 완전히 동일해 재사용했다(파라미터 이름만 rows/metric_rows로 달랐음).
# ===========================================================================

OUTPUT_DIR = "example/result/공문자료"


def find_months_needing_final_doc(master_path, output_dir):
    """검증·지표가 모두 끝났지만 아직 최종 md가 없는 달을 찾는다."""
    wb = openpyxl.load_workbook(master_path)
    sheetnames = set(wb.sheetnames)

    targets = []
    for name in wb.sheetnames:
        match = METRIC_SHEET_PATTERN.match(name)
        if not match:
            continue
        sheet_name = match.group(1)

        if f"검증로그_{sheet_name}" not in sheetnames:
            print(f"건너뜀 (검증 미완료): {sheet_name}")
            continue

        out_path = os.path.join(output_dir, f"공문작성자료_{sheet_name}.md")
        if os.path.exists(out_path):
            print(f"건너뜀 (이미 생성됨): {sheet_name} → {out_path} 존재")
        else:
            print(f"신규 발견: {sheet_name} → 최종 자료 생성 필요")
            targets.append(sheet_name)

    return targets


def build_rate_interpretation(agg):
    """집행률 그래프와 동일한 집계로 해석 문장을 만든다 (이미지 분석이 아니라 데이터 기반)."""
    items = []
    for (사업장, 부서명), v in agg.items():
        rate = v["집행액"] / v["배정액"] * 100 if v["배정액"] else 0
        items.append((사업장, 부서명, rate, v["미확정"]))

    avg = sum(i[2] for i in items) / len(items)
    best = max(items, key=lambda i: i[2])
    worst = min(items, key=lambda i: i[2])

    lines = [
        f"- 부서 평균 집행률은 {avg:.1f}%이며, {worst[2]:.1f}%~{best[2]:.1f}% 범위에 분포합니다.",
        f"- 집행률이 가장 높은 부서는 {best[0]} {best[1]}({best[2]:.1f}%), "
        f"가장 낮은 부서는 {worst[0]} {worst[1]}({worst[2]:.1f}%)입니다.",
    ]

    partial = [f"{s} {d}" for s, d, _, p in items if p]
    if partial:
        lines.append(
            f"- {', '.join(partial)}는 일부 항목의 집행액이 미확정 상태라 "
            f"집행률이 실제보다 낮게 나타났을 수 있습니다 (그래프에 빗금으로 표시됨)."
        )

    return "\n".join(lines)


def build_delta_interpretation(agg):
    """증감 그래프와 동일한 집계로 해석 문장을 만든다."""
    items = [
        (사업장, 부서명, v["증감합"], v["미확정"])
        for (사업장, 부서명), v in agg.items()
        if v["증감건수"] > 0
    ]
    if not items:
        return None

    best = max(items, key=lambda i: i[2])
    worst = min(items, key=lambda i: i[2])

    if all(i[2] >= 0 for i in items):
        trend = "모든 부서에서 전월 대비 집행액이 증가했습니다."
    else:
        trend = "일부 부서에서 전월 대비 집행액이 감소했습니다."

    lines = [
        f"- {trend}",
        f"- 증가 폭이 가장 큰 부서는 {best[0]} {best[1]}({best[2]:+,.0f}원), "
        f"가장 작은 부서는 {worst[0]} {worst[1]}({worst[2]:+,.0f}원)입니다.",
    ]
    return "\n".join(lines)


def find_reply_files(email_dir, sheet_name):
    """'이메일_응답_*_연도-월.txt' 패턴의 담당자 회신 파일을 전부 찾는다."""
    if not os.path.isdir(email_dir):
        return []
    pattern = os.path.join(email_dir, f"이메일_응답_*_{sheet_name}.txt")
    return sorted(glob.glob(pattern))


def build_final_markdown(master_path, graph_dir, email_dir, sheet_name):
    metric_rows = load_merged_rows(master_path, f"지표_{sheet_name}")
    log_rows = load_merged_rows(master_path, f"검증로그_{sheet_name}")
    agg = aggregate_by_dept(metric_rows)

    미제출부서 = [r["내용"] for r in log_rows if r["구분"] == "미제출부서"]
    누락값 = [r["내용"] for r in log_rows if r["구분"] == "누락값"]
    이상치 = [r["내용"] for r in log_rows if r["구분"] == "이상치"]

    rate_img = os.path.join(graph_dir, f"집행률_비교_{sheet_name}.png")
    delta_img = os.path.join(graph_dir, f"증감_비교_{sheet_name}.png")

    lines = [f"# {sheet_name} 예산 집행 현황 — 공문 초안 작성 자료\n"]
    lines.append(
        "> 이 문서는 데이터 검증, 지표 계산, 비교 그래프, 담당자 회신을 통합해 "
        "자동 생성되었습니다. 공문 초안 작성 시 이 문서를 참고 자료로 사용하세요.\n"
    )

    lines.append("## 1. 부서별 집행 현황 요약\n")
    lines.append("| 사업장 | 부서 | 배정액 | 집행액 | 집행률 | 비고 |")
    lines.append("|---|---|---:|---:|---:|---|")
    for (사업장, 부서명), v in agg.items():
        rate = v["집행액"] / v["배정액"] * 100 if v["배정액"] else 0
        비고 = "일부 항목 미확정" if v["미확정"] else ""
        lines.append(
            f"| {사업장} | {부서명} | {v['배정액']:,}원 | {v['집행액']:,}원 | {rate:.1f}% | {비고} |"
        )
    lines.append("")

    lines.append("## 2. 첨부 그래프\n")

    lines.append("### 그래프 1. 부서별 집행률 비교")
    if os.path.exists(rate_img):
        lines.append(f"- 파일 위치: `{rate_img}`")
        lines.append("- 해석:")
        lines.append(build_rate_interpretation(agg))
    else:
        lines.append("- (그래프가 아직 생성되지 않았습니다. Step 7을 먼저 실행하세요.)")
    lines.append("")

    lines.append("### 그래프 2. 부서별 전월 대비 집행액 증감")
    if os.path.exists(delta_img):
        interpretation = build_delta_interpretation(agg)
        lines.append(f"- 파일 위치: `{delta_img}`")
        lines.append("- 해석:")
        lines.append(interpretation)
    else:
        lines.append("- 전월 데이터가 없어 이번 달은 생성되지 않았습니다.")
    lines.append("")

    lines.append("## 3. 확인 필요 항목 (검증 결과)\n")
    if not (미제출부서 or 누락값 or 이상치):
        lines.append("- 검증 결과 이상 없음.\n")
    else:
        if 미제출부서:
            lines.append("**미제출 부서**")
            lines.extend(f"- {x}" for x in 미제출부서)
        if 누락값:
            lines.append("**누락값**")
            lines.extend(f"- {x}" for x in 누락값)
        if 이상치:
            lines.append("**이상치**")
            lines.extend(f"- {x}" for x in 이상치)
        lines.append("")

    lines.append("## 4. 담당자 회신 내용\n")
    reply_paths = find_reply_files(email_dir, sheet_name)
    if not reply_paths:
        lines.append("- 접수된 담당자 회신이 없습니다.\n")
    else:
        for path in reply_paths:
            with open(path, encoding="utf-8") as f:
                content = f.read().strip()
            lines.append(f"### 회신 파일: `{path}`\n")
            lines.append("```")
            lines.append(content)
            lines.append("```\n")

    return "\n".join(lines)


def save_final_markdown(content, output_dir, sheet_name):
    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, f"공문작성자료_{sheet_name}.md")
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"저장 완료: {path}")
    return path


def run_final_doc_pipeline(master_path=MASTER_FILE, graph_dir=GRAPH_DIR, email_dir=EMAIL_DIR, output_dir=OUTPUT_DIR):
    targets = find_months_needing_final_doc(master_path, output_dir)

    for sheet_name in targets:
        content = build_final_markdown(master_path, graph_dir, email_dir, sheet_name)
        save_final_markdown(content, output_dir, sheet_name)


# ===========================================================================
# Step 8-2 — 보고서 템플릿 준비 (vibecoding.ipynb 코드 셀 그대로, 목차·작성 가이드)
# ===========================================================================

TEMPLATE_PATH = "example/result/월간예산집행보고서_템플릿.md"

TEMPLATE_CONTENT = """# 월간 예산 집행 현황 보고서 — 목차 및 작성 가이드

> 보고 대상: **내부 경영진(원장/이사장 및 관련 임원)**
> 성격: 결재용 공문이 아니라 **의사결정 지원용 월례 보고서**
> 핵심 질문(이 보고서가 답해야 하는 것): "우리 조직은 계획대로 예산을 쓰고 있는가? 문제가 있는 부서는 어디이고 왜인가? 지금 결정해야 할 것이 있는가?"

이 문서는 매달 반복 작성되는 보고서의 **목차와 작성 기준**을 고정해 둔 템플릿입니다. Step 9에서 Upstage API로 보고서 초안을 생성할 때, 이 템플릿의 목차·관점·톤을 그대로 따르도록 프롬프트에 포함시킵니다.

---

## 0. 문서 제목 (표지 역할)

**형식**: `{연도}년 {월}월 예산 집행 현황 보고`

- 부제목으로 보고 대상 기간과 작성 부서를 한 줄로 명시: `(보고기간: 2026.07.01~2026.07.31 / 작성: 재무회계부)`
- 경영진이 문서함에서 스캔할 때 월과 핵심 상태(정상/이상)를 즉시 알 수 있도록, 이상 부서가 있는 달은 제목 옆에 `[확인 필요 1건]` 같은 태그를 붙이는 것을 권장.

---

## 1. 핵심 요약 (Executive Summary)

**목적**: 경영진은 전체 문서를 읽지 않고 이 섹션만 봐도 판단이 서야 합니다. **결론이 맨 위**, 근거는 그 다음입니다.

**관점**: "우리 조직 전체"가 단위입니다. 부서 하나하나가 아니라 기관 전체 집행률·추세를 먼저 말하고, 그 다음에 예외를 언급합니다.

**작성 방식**:
- 3~4줄 이내. 문장형이 아니라 **불릿 3개**로: ① 전체 집행률(전월 대비 증감 포함) ② 정상/이상 부서 개수 ③ 지금 경영진이 알아야 할 결정 사항이 있는지 여부
- 숫자는 있는 그대로 쓰되, "목표 대비"라는 기준이 있다면 반드시 그 기준과 비교해서 서술 (없으면 "5개 부서 평균 대비"로 대체)
- 이상 부서가 있다면 여기서 이름과 상태를 먼저 언급하고, 상세 사유는 3번 섹션으로 미룹니다.

**데이터 소스**: `공문자료/공문작성자료_{연도-월}.md`의 "부서별 집행 현황 요약" 표 전체 집계, "확인 필요 항목" 존재 여부.

---

## 2. 부서별 집행 현황

**목적**: 경영진이 "어느 부서가 잘하고 있고, 어느 부서가 뒤처지는지"를 한눈에 비교하게 합니다.

**관점**: 개별 부서 실적 나열이 아니라 **비교**가 핵심입니다. 평균과 대비해서 위/아래를 명확히 구분해 서술합니다.

**작성 방식**:
- 표 하나(부서/배정액/집행액/집행률/비고)로 시작 — `지표_{연도-월}` 집계 결과 그대로 사용
- 표 아래에 그래프 2종을 첨부하고, **파일 경로를 정확히 명시**: `그래프/집행률_비교_{연도-월}.png`, `그래프/증감_비교_{연도-월}.png`
- 그래프 각각에 1~2줄 캡션(해석)을 붙입니다. 캡션은 "그래프를 보면 안다"가 아니라 "그래프를 안 봐도 요점이 전달된다"를 목표로 씁니다 (예: "안전경영부만 평균보다 크게 낮은데, 이는 인건비 항목 미확정 때문이며 실질 집행률은 정상 범위임").
- 미확정/이상 항목이 섞인 부서는 반드시 각주나 비고로 "왜 낮게/이상하게 보이는지"를 짚어줘야 합니다. 안 그러면 경영진이 실제보다 나쁘게 오해합니다.

**데이터 소스**: `공문작성자료_{연도-월}.md`의 표 + 그래프 해석 문장을 거의 그대로 재사용 가능.

---

## 3. 확인 필요 항목 및 처리 결과

**목적**: 이상 신호(누락값/이상치/미제출)가 "왜 발생했고, 지금 어떤 상태인지"를 경영진에게 설명해 불필요한 우려를 없애거나, 반대로 진짜 조치가 필요한 건 명확히 알립니다.

**관점**: 실무자(재무회계부) 관점이 아니라 **경영진이 판단할 수 있는 수준의 요약**이어야 합니다. "왜 이상치가 떴는지"의 기술적 디테일(자릿수 오류 등)은 한 줄로 압축하고, "그래서 지금 문제가 있는지 없는지"를 명확히 씁니다.

**작성 방식**: 항목별로 아래 3단 구조를 반복합니다.
1. **무엇이 발견됐나** (검증 결과 원문 요약, 1줄)
2. **확인된 사유** (담당자 회신 기반, 1~2줄 — 회신 원문을 그대로 붙이지 말고 요지만)
3. **현재 상태** — "정정 완료" / "다음 달 반영 예정" / "추가 확인 필요" 중 하나로 명확히 라벨링

이상이 하나도 없는 달은 이 섹션을 짧게 "이번 달 검증 결과 특이사항 없음"으로 끝내고 다음 섹션으로 넘어갑니다 (경영진 보고서에서 "문제 없음"도 중요한 정보입니다).

**데이터 소스**: `검증로그_{연도-월}` 시트(누락값/이상치/미제출부서) + `이메일_응답_*_{연도-월}.txt`(담당자 회신 요지).

---

## 4. 차월 계획 및 전망

**목적**: 경영진이 "다음 달에 뭘 지켜봐야 하는지" 미리 알게 합니다. 월례 보고서를 시계열로 이어주는 섹션입니다.

**관점**: 예측이 아니라 **이미 확정되었거나 담당자가 언급한 계획**만 씁니다. 근거 없는 낙관/비관 전망은 배제합니다.

**작성 방식**:
- 담당자 회신에 차월 계획 언급이 있으면 그대로 인용 (예: "하반기 안전점검 위탁용역 계약 체결 예정 → 사업비 집행률 증가 전망")
- 특별한 계획 언급이 없는 부서는 "통상적인 집행 추세 지속 전망"으로 간단히 처리
- 이번 달 미해결 항목(예: 인건비 소급 반영 예정)이 있다면 여기서 "다음 보고서에서 확인할 항목"으로 명시해, 다음 달 보고서와 자연스럽게 연결

**데이터 소스**: 담당자 회신 텍스트 중 계획/향후 일정 언급 부분.

---

## 5. 경영진 확인·결정 요청 사항 (해당하는 경우에만)

**목적**: 경영진이 실제로 뭔가를 결정하거나 승인해야 하는 경우에만 넣는 섹션입니다. 없는 달에는 섹션 자체를 생략합니다.

**관점**: "정보 전달"이 아니라 "행동 요청"입니다. 무엇을, 왜, 언제까지 결정해야 하는지 명확히.

**작성 방식**: 있다면 불릿 1~3개로 간결하게. 예: "정정 자료 승인 필요(제출 기한 8/10)" 같이 기한과 함께.

**데이터 소스**: 담당자 회신 중 승인/의사결정이 필요하다고 판단되는 내용 (자동 추출이 어려우면 재무회계부 담당자가 수동으로 판단해 추가).

---

## 전체 톤 가이드

- **결론 우선, 디테일은 뒤로**: 경영진 보고서는 위에서 아래로 읽다가 아무 데서나 멈춰도 핵심이 전달돼야 합니다.
- **숫자는 비교와 함께**: "60.5%"보다 "평균보다 7%p 높음"이 더 유용합니다.
- **이상 신호는 숨기지 않되 과장하지 않음**: 사유가 확인된 이상치는 "이상치였으나 확인 결과 정상"으로 명확히 마무리 짓습니다. 애매하게 남겨두지 않습니다.
- **문장은 짧게, 격식은 유지**: 내부 보고서이지만 경영진 대상이므로 구어체·이모지는 쓰지 않고, 존댓말과 개조식을 혼용합니다.
- **분량**: 전체 A4 1~2장 분량을 넘기지 않습니다 (첨부 그래프 제외). 길어지면 경영진이 안 읽습니다.
"""


def create_report_template(path=TEMPLATE_PATH, content=TEMPLATE_CONTENT):
    """보고서 목차·작성 가이드 템플릿을 만든다. 이미 있으면 건너뛴다 (한 번 만들면 계속 재사용)."""
    if os.path.exists(path):
        print(f"건너뜀 (이미 있음): {path}")
        return

    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"저장 완료: {path}")


# ===========================================================================
# Step 9 — 보고서 초안 생성 (vibecoding.ipynb 코드 셀 그대로, LLM · Upstage API)
#
# 노트북 셀은 여기서도 OpenAI 클라이언트를 새로 만드는데, Step 3-1에서 만든
# _upstage_client를 그대로 재사용했다(클라이언트를 두 개 둘 필요가 없다).
# ===========================================================================

DRAFT_DIR = "example/result/보고서초안"
DATA_PATTERN = re.compile(r"^공문작성자료_(\d{4}-\d{2})\.md$")


def find_months_needing_draft(data_dir, draft_dir):
    """원자료는 있지만 아직 보고서 초안이 없는 달을 찾는다."""
    targets = []
    if not os.path.isdir(data_dir):
        return targets

    for filename in sorted(os.listdir(data_dir)):
        match = DATA_PATTERN.match(filename)
        if not match:
            continue

        sheet_name = match.group(1)
        draft_path = os.path.join(draft_dir, f"보고서초안_{sheet_name}.md")
        if os.path.exists(draft_path):
            print(f"건너뜀 (이미 생성됨): {sheet_name} → {draft_path} 존재")
        else:
            print(f"신규 발견: {sheet_name} → 보고서 초안 생성 필요")
            targets.append(sheet_name)

    return targets


def load_text(path):
    with open(path, encoding="utf-8") as f:
        return f.read()


def draft_report(sheet_name, template_text, data_text):
    """Upstage Solar Pro에게 템플릿 기준으로 보고서 초안을 요청한다."""
    prompt = f"""아래 [작성 가이드]의 목차와 기준을 반드시 따라서, [원자료]를 바탕으로 {sheet_name} 월간 예산 집행 현황 보고서를 작성해주세요.

[작성 가이드]
{template_text}

[원자료]
{data_text}

요구사항:
- [작성 가이드]의 목차 순서와 각 섹션의 관점·톤을 그대로 따라주세요.
- 그래프는 마크다운 이미지 문법으로 정확한 경로를 넣어주세요 (예: ![집행률 비교](그래프/집행률_비교_{sheet_name}.png)).
- [원자료]에 없는 내용은 지어내지 말고, 있는 정보만 사용해주세요. 특히 [원자료]에 나오지 않는 비교 수치(예: 전월 평균 %, %p 증감 등)는 절대로 계산하거나 추정해서 만들어내지 마세요. 그런 수치가 근거자료에 없다면 "증가 추세" 같은 정성적 표현만 사용하세요.
- 완성된 보고서 전체를 markdown으로 출력해주세요 (다른 설명 없이 보고서 본문만).
"""
    response = _upstage_client.chat.completions.create(
        model="solar-pro",
        messages=[{"role": "user", "content": prompt}],
    )
    return response.choices[0].message.content


def save_draft(content, draft_dir, sheet_name):
    os.makedirs(draft_dir, exist_ok=True)
    path = os.path.join(draft_dir, f"보고서초안_{sheet_name}.md")
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"저장 완료: {path}")
    return path


def run_draft_pipeline(data_dir=OUTPUT_DIR, template_path=TEMPLATE_PATH, draft_dir=DRAFT_DIR):
    template_text = load_text(template_path)
    targets = find_months_needing_draft(data_dir, draft_dir)

    for sheet_name in targets:
        data_path = os.path.join(data_dir, f"공문작성자료_{sheet_name}.md")
        data_text = load_text(data_path)

        print(f"\n[{sheet_name}] 보고서 초안 생성 중...")
        content = draft_report(sheet_name, template_text, data_text)
        save_draft(content, draft_dir, sheet_name)
